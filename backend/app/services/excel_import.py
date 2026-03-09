"""Excel 导入服务。"""

import re
from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.enums import Gender, Source
from app.models.person import Person


# Excel 列映射
COLUMN_MAPPING = {
    "A": "name",
    "B": "phone",
    "C": "city",
    "D": "id_card",
    "E": "gender",
    "F": "age",
    "G": "source",
    "H": "address",
    "I": "remark",
    "J": "reusable",
}

# 中文到枚举的映射
GENDER_MAP = {
    "男": Gender.MALE,
    "女": Gender.FEMALE,
    "未知": Gender.UNKNOWN,
}

SOURCE_MAP = {
    "BOSS直聘": Source.BOSS,
    "快手": Source.KUAISHOU,
    "抖音": Source.DOUYIN,
    "58同城": Source.FIVE_EIGHT,
    "内推": Source.REFERRAL,
    "其他": Source.OTHER,
}

# 手机号正则
PHONE_PATTERN = re.compile(r"^1[3-9]\d{9}$")

# 最大行数
MAX_IMPORT_ROWS = 1000


class ExcelImportService:
    """Excel 导入服务类。"""

    def generate_template(self) -> BytesIO:
        """生成导入模板。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "人员导入模板"

        # 表头
        headers = [
            "姓名",
            "手机号",
            "城市",
            "身份证号",
            "性别",
            "年龄",
            "来源渠道",
            "详细地址",
            "备注",
            "可复用",
        ]
        ws.append(headers)

        # 示例数据
        ws.append([
            "张三",
            "13800138000",
            "北京",
            "110101199001011234",
            "男",
            "28",
            "BOSS直聘",
            "朝阳区xx路xx号",
            "有快递经验",
            "是",
        ])

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def import_persons(
        self,
        session: AsyncSession,
        file_content: bytes,
        created_by: int,
    ) -> dict:
        """导入人员数据。"""
        errors: list[dict] = []
        success_count = 0

        # 加载 Excel
        try:
            wb = load_workbook(BytesIO(file_content))
            ws = wb.active
        except Exception as e:
            return {
                "success_count": 0,
                "fail_count": 0,
                "total": 0,
                "errors": [{"row": 0, "field": None, "message": f"文件解析失败: {str(e)}"}],
            }

        # 获取数据行（跳过表头）
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total = len(rows)

        # 检查行数限制
        if total > MAX_IMPORT_ROWS:
            return {
                "success_count": 0,
                "fail_count": 0,
                "total": total,
                "errors": [{"row": 0, "field": None, "message": f"单次最多导入{MAX_IMPORT_ROWS}行"}],
            }

        # 获取已存在的手机号
        phones_in_file = [self._get_value(row, 1) for row in rows if self._get_value(row, 1)]
        existing_phones = await self._get_existing_phones(session, phones_in_file)

        # 逐行处理
        for idx, row in enumerate(rows, start=2):  # start=2 因为第1行是表头
            try:
                person_data = self._parse_row(row, idx, existing_phones)
                if person_data is None:
                    continue  # 空行跳过

                # 创建人员
                person = Person(
                    **person_data,
                    created_by=created_by,
                )
                session.add(person)
                success_count += 1

                # 记录已处理的手机号（防止同文件内重复）
                existing_phones.add(person_data["phone"])

            except ValueError as e:
                errors.append({"row": idx, "field": None, "message": str(e)})

        # 提交事务
        if success_count > 0:
            await session.commit()

        return {
            "success_count": success_count,
            "fail_count": len(errors),
            "total": total,
            "errors": errors,
        }

    def _get_value(self, row: tuple, index: int) -> str | None:
        """获取行中的值。"""
        if index >= len(row):
            return None
        value = row[index]
        if value is None:
            return None
        return str(value).strip()

    def _parse_row(self, row: tuple, row_num: int, existing_phones: set) -> dict | None:
        """解析行数据。"""
        # 获取各字段值
        name = self._get_value(row, 0)
        phone = self._get_value(row, 1)
        city = self._get_value(row, 2)
        id_card = self._get_value(row, 3)
        gender_str = self._get_value(row, 4)
        age_str = self._get_value(row, 5)
        source_str = self._get_value(row, 6)
        address = self._get_value(row, 7)
        remark = self._get_value(row, 8)
        reusable_str = self._get_value(row, 9)

        # 空行检查
        if not name and not phone and not city:
            return None

        # 必填字段校验
        if not name:
            raise ValueError("姓名不能为空")
        if not phone:
            raise ValueError("手机号不能为空")
        if not city:
            raise ValueError("城市不能为空")

        # 手机号格式校验
        if not PHONE_PATTERN.match(phone):
            raise ValueError(f"手机号格式错误: {phone}")

        # 手机号重复校验
        if phone in existing_phones:
            raise ValueError(f"手机号已存在: {phone}")

        # 性别转换
        gender = GENDER_MAP.get(gender_str or "未知", Gender.UNKNOWN)

        # 年龄转换
        age = None
        if age_str:
            try:
                age = int(age_str)
                if age < 0 or age > 150:
                    raise ValueError("年龄必须在0-150之间")
            except ValueError:
                raise ValueError(f"年龄格式错误: {age_str}")

        # 来源转换
        source = SOURCE_MAP.get(source_str or "其他", Source.OTHER)

        # 可复用转换
        reusable = True
        if reusable_str:
            reusable = reusable_str.lower() in ("是", "yes", "true", "1")

        return {
            "name": name,
            "phone": phone,
            "city": city,
            "id_card": id_card,
            "gender": gender,
            "age": age,
            "source": source,
            "address": address,
            "remark": remark,
            "reusable": reusable,
        }

    async def _get_existing_phones(
        self, session: AsyncSession, phones: list[str]
    ) -> set[str]:
        """获取已存在的手机号。"""
        if not phones:
            return set()

        result = await session.execute(
            select(Person.phone).where(Person.phone.in_(phones))
        )
        return {row[0] for row in result.all()}


# 单例
excel_import_service = ExcelImportService()

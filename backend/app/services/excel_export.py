"""Excel 导出服务。"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.models.enums import Gender, PersonProjectStatus, Source
from app.models.person import Person
from app.models.person_project import PersonProject
from app.models.project import Project


# 状态中文映射
STATUS_MAP = {
    PersonProjectStatus.SIGNED_UP: "已报名",
    PersonProjectStatus.INVITED: "已邀约",
    PersonProjectStatus.INTERVIEW_PENDING: "待面试",
    PersonProjectStatus.INTERVIEWED: "已面试",
    PersonProjectStatus.IN_TRIAL: "试工中",
    PersonProjectStatus.TRIAL_PASSED: "试工通过",
    PersonProjectStatus.FAILED: "失败",
    PersonProjectStatus.UNREACHABLE: "联系不上",
}

# 性别中文映射
GENDER_DISPLAY_MAP = {
    Gender.MALE: "男",
    Gender.FEMALE: "女",
    Gender.UNKNOWN: "未知",
}

# 来源中文映射
SOURCE_DISPLAY_MAP = {
    Source.BOSS: "BOSS直聘",
    Source.KUAISHOU: "快手",
    Source.DOUYIN: "抖音",
    Source.FIVE_EIGHT: "58同城",
    Source.REFERRAL: "内推",
    Source.OTHER: "其他",
}


class ExcelExportService:
    """Excel 导出服务类。"""

    def _create_header_style(self) -> dict:
        """创建表头样式。"""
        return {
            "font": Font(bold=True),
            "fill": PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        }

    def _apply_header_style(self, ws, row_num: int = 1):
        """应用表头样式。"""
        style = self._create_header_style()
        for cell in ws[row_num]:
            for key, value in style.items():
                setattr(cell, key, value)

    async def export_persons(
        self,
        session: AsyncSession,
        project_id: int | None = None,
        status_filter: PersonProjectStatus | None = None,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
    ) -> BytesIO:
        """导出人员列表。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "人员列表"

        # 表头
        headers = [
            "ID",
            "姓名",
            "手机号",
            "城市",
            "身份证号",
            "性别",
            "年龄",
            "来源渠道",
            "详细地址",
            "备注",
            "可复用",
            "项目",
            "状态",
            "负责人ID",
            "创建时间",
        ]
        ws.append(headers)
        self._apply_header_style(ws)

        # 查询数据
        query = (
            select(Person, PersonProject, Project)
            .outerjoin(PersonProject, Person.id == PersonProject.person_id)
            .outerjoin(Project, PersonProject.project_id == Project.id)
        )

        if project_id:
            query = query.where(PersonProject.project_id == project_id)
        if status_filter:
            query = query.where(PersonProject.status == status_filter)
        if start_date:
            query = query.where(Person.created_at >= start_date)
        if end_date:
            query = query.where(Person.created_at < end_date)

        query = query.order_by(Person.created_at.desc())

        result = await session.execute(query)
        rows = result.all()

        # 填充数据
        for person, pp, project in rows:
            ws.append([
                person.id,
                person.name,
                person.phone,
                person.city,
                person.id_card,
                GENDER_DISPLAY_MAP.get(person.gender, "未知"),
                person.age,
                SOURCE_DISPLAY_MAP.get(person.source, "其他"),
                person.address,
                person.remark,
                "是" if person.reusable else "否",
                project.name if project else "",
                STATUS_MAP.get(pp.status, "") if pp else "",
                pp.owner_id if pp else "",
                person.created_at.strftime("%Y-%m-%d %H:%M:%S") if person.created_at else "",
            ])

        # 调整列宽
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def export_statistics(
        self,
        session: AsyncSession,
        project_id: int,
        start_date: datetime,
        end_date: datetime,
        funnel_data: dict,
        daily_data: dict,
        fail_reasons: dict,
    ) -> BytesIO:
        """导出统计数据。"""
        wb = Workbook()

        # 获取项目信息
        project_result = await session.execute(
            select(Project).where(Project.id == project_id)
        )
        project = project_result.scalar_one_or_none()
        project_name = project.name if project else f"项目{project_id}"

        # Sheet 1: 漏斗数据
        ws1 = wb.active
        ws1.title = "漏斗数据"

        ws1.append(["项目统计漏斗"])
        ws1.append([f"项目: {project_name}"])
        ws1.append([f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws1.append([])
        ws1.append(["状态", "人数", "占比"])
        self._apply_header_style(ws1, 5)

        total = funnel_data.get("total", 0)
        status_counts = funnel_data.get("status_counts", {})

        for status_value, count in status_counts.items():
            status_name = STATUS_MAP.get(PersonProjectStatus(status_value), status_value)
            rate = f"{count / total * 100:.2f}%" if total > 0 else "0%"
            ws1.append([status_name, count, rate])

        ws1.append([])
        ws1.append(["总计", total, "100%"])

        # Sheet 2: 每日统计
        ws2 = wb.create_sheet("每日统计")
        ws2.append(["日期", "新增人数"])
        self._apply_header_style(ws2)

        for date_str, count in daily_data.get("daily_stats", {}).items():
            ws2.append([date_str, count])

        # Sheet 3: 失败原因
        ws3 = wb.create_sheet("失败原因")
        ws3.append(["失败原因", "人数", "占比"])
        self._apply_header_style(ws3)

        fail_total = sum(fail_reasons.values()) if fail_reasons else 0
        fail_reason_map = {
            "no_show": "爽约",
            "rejected": "拒绝",
            "no_package": "未购包",
            "no_training": "未到训",
            "rescheduled": "改约",
            "abandoned": "放弃",
            "not_qualified": "不符合要求",
            "other": "其他",
        }

        for reason_value, count in fail_reasons.items():
            reason_name = fail_reason_map.get(reason_value, reason_value)
            rate = f"{count / fail_total * 100:.2f}%" if fail_total > 0 else "0%"
            ws3.append([reason_name, count, rate])

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# 单例
excel_export_service = ExcelExportService()
